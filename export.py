import statistics
from matplotlib.patches import Patch
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from constants import *
from sim import calcola_perdite_totali

def _font(b=False, s=11, c="000000"):
    return Font(name="Calibri", bold=b, size=s, color=c)

def _bordo():
    lato = Side(style="thin", color="BFBFBF")
    return Border(left=lato, right=lato, top=lato, bottom=lato)

def _allinea(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def _adatta_colonne(ws, min_w=12, max_w=50):
    for col in ws.columns:
        larghezza = min_w
        for cell in col:
            if cell.value:
                larghezza = max(larghezza, min(len(str(cell.value)) + 4, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = larghezza

def _scrivi_riga(ws, dati, f_style, bg_color=None, align="left", merge_cols=None, height=15):
    ws.append(dati)
    riga = ws.max_row
    if merge_cols:
        ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=merge_cols)
    
    for i in range(1, (merge_cols or len(dati)) + 1):
        c = ws.cell(riga, i)
        c.font = f_style
        if bg_color:
            c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = _allinea("center" if merge_cols else align)
        c.border = _bordo()
    ws.row_dimensions[riga].height = height

def _riga_titolo(ws, testo, n_colonne, colore="1F3864"):
    _scrivi_riga(ws, [testo], _font(True, 13, "FFFFFF"), colore, merge_cols=n_colonne, height=22)

def _riga_intestazione(ws, valori, colore="2E75B6"):
    _scrivi_riga(ws, valori, _font(True, 11, "FFFFFF"), colore, align="center", height=18)

def _riga_dati(ws, valori, pari=True, evidenzia=False):
    sfondo = "FFF2CC" if evidenzia else ("EBF3FB" if pari else "FFFFFF")
    _scrivi_riga(ws, valori, _font(), sfondo)

def _riga_vuota(ws, n_colonne):
    ws.append([""] * n_colonne)
    ws.row_dimensions[ws.max_row].height = 8

def _xlsx_risultato_sintetico(r, ws):
    quantita_tot = round(sum(r["quantita"][c] for c in COLTURE), 2)
    perdite_tot = calcola_perdite_totali(r)
    
    _riga_titolo(ws, "RISULTATO FINALE DEL LOTTO DI PRODUZIONE", 2)
    _riga_intestazione(ws, ["Indicatore", "Valore"])
    
    dati = [
        ("Tempo totale processo (ore)", r["t_totale"]),
        ("Tempo totale processo (giorni lav.)", r["t_totale_gg"]),
        ("Perdite totali (t)", perdite_tot),
        ("Perdite totali (% sul raccolto)", f"{round(perdite_tot / quantita_tot * 100, 2)}%"),
    ]
    
    for i, (label, val) in enumerate(dati):
        _riga_dati(ws, [label, val], pari=(i % 2 == 0))
    _adatta_colonne(ws)

def _xlsx_dettaglio_completo(r, ws, origine):
    N = 9
    _riga_titolo(ws, f"PARAMETRI DELLA SIMULAZIONE ({origine})", 2)
    _riga_intestazione(ws, ["Parametro", "Valore"])
    
    macchine = '1 mietitrebbia' if r['scenario_raccolta'] == 'standard' else '2 mietitrebbie'
    params = [
        ("Scenario raccolta", f"{r['scenario_raccolta'].capitalize()}  ({macchine})"),
        ("Annata climatica", r["scenario_climatico"].capitalize()),
    ]
    
    for c in COLTURE:
        resa = r["scenario_resa"][c]
        t_ha = SCENARI_RESA_RACCOLTA[c][resa]
        params.append((f"Resa {c.capitalize()}", f"{resa.capitalize()}  ({t_ha} t/ha  -  {r['quantita'][c]:.2f} t)"))
        
    for i, (label, val) in enumerate(params):
        _riga_dati(ws, [label, val], pari=(i % 2 == 0))

    _riga_vuota(ws, N)
    _riga_titolo(ws, "DETTAGLIO PER FASE E COLTURA", N)
    _riga_intestazione(ws, ["Coltura", "Fase", "Linea", "Durata (h)", "Durata (gg)", "Attesa coda (h)", "Perdita (t)", "Sacchi prodotti", "Note"])

    colori_coltura = {"mais": "FFF9C4", "girasole": "E8F5E9", "soia": "E3F2FD"}

    for c in COLTURE:
        f = r["fasi"][c]
        es = f["essiccazione"]
        nota_es = f"SALTATA ({UMIDITA_INIZIALE[c][r['scenario_climatico']]}% < {UMIDITA_TARGET[c]}%)" if es["tempo_ore"] == 0 else "-"
        
        dati_fasi = [
            ["Raccolta", "-", f["raccolta"]["tempo_ore"], f["raccolta"]["tempo_gg"], "-", "-", "-", f"Cap. {f['raccolta']['cap_oraria']} t/h"],
            ["Pulizia", "-", f["pulizia"]["tempo_ore"], f["pulizia"]["tempo_gg"], f["pulizia"]["t_attesa"], f["pulizia"]["perdita"], "-", "-"],
            ["Essiccazione", "-", es.get("tempo_ore", "-"), es.get("tempo_gg", "-"), es.get("t_attesa", "-"), es.get("perdita", "-"), "-", nota_es],
            ["Stoccaggio", "-", "trascurabile", "-", "-", f["stoccaggio"]["perdita"], "-", "-"],
            ["Divisione Linee", "A / B", "-", "-", "-", "-", "-", f"A {f['split']['perc_A']}%  B {f['split']['perc_B']}%"],
            ["Confezionamento", "A", f["confez_A"]["tempo_ore"], f["confez_A"]["tempo_gg"], f["confez_A"]["t_attesa"], "-", f"{f['confez_A']['n_sacchi']:,}", NOMI_PRODOTTI[f"{c}_A"]],
        ]
        
        if c == "soia":
            dati_fasi.append(["Tostatura", "B", f["tostatura"]["tempo_ore"], f["tostatura"]["tempo_gg"], "-", f["tostatura"]["perdita"], "-", "Solo Soia"])
            
        dati_fasi.extend([
            ["Macinazione", "B", f["macinazione"]["tempo_ore"], f["macinazione"]["tempo_gg"], f["macinazione"]["t_attesa"], f["macinazione"]["perdita"], "-", NOMI_PRODOTTI[f"{c}_B"]],
            ["Confez. Farina", "B", f["confez_B"]["tempo_ore"], f["confez_B"]["tempo_gg"], f["confez_B"]["t_attesa"], "-", f"{f['confez_B']['n_sacchi']:,}", NOMI_PRODOTTI[f"{c}_B"]]
        ])

        for riga in dati_fasi:
            _scrivi_riga(ws, [c.capitalize()] + riga, _font(), colori_coltura[c])
            
        _riga_vuota(ws, N)
    _adatta_colonne(ws)

def _xlsx_simulazione(r, nome_file, origine):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Risultato Finale"
    _xlsx_risultato_sintetico(r, ws1)
    
    ws2 = wb.create_sheet("Dettaglio Fasi")
    _xlsx_dettaglio_completo(r, ws2, origine)
    
    wb.save(nome_file)

def esporta_excel_singola_random(r, nome_file="simulazione_random.xlsx"):
    _xlsx_simulazione(r, nome_file, "parametri estratti casualmente")

def esporta_excel_manuale(r, nome_file="simulazione_manuale.xlsx"):
    _xlsx_simulazione(r, nome_file, "parametri configurati manualmente")

def esporta_excel_seed(r, nome_file):
    _xlsx_simulazione(r, nome_file, "parametri estratti con seed fisso")

def _stile_plot(ax, titolo, xlabel="", ylabel=""):
    ax.set_title(titolo, fontweight="bold")
    if xlabel: ax.set_xlabel(xlabel)
    if ylabel: ax.set_ylabel(ylabel)
    ax.grid(axis="y" if ylabel else "x", color="#E0E0E0", linewidth=0.8)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

def crea_figure_analisi(tutti_i_risultati):
    plt.rcParams.update({"font.family": "DejaVu Sans", "axes.titlesize": 13, "axes.labelsize": 11, "axes.titlepad": 14})
    BLU, ARANCIO, ROSSO = "#2E75B6", "#E8A838", "#C00000"
    
    tempi = [r["t_totale"] for r in tutti_i_risultati]
    media, dev = statistics.mean(tempi), statistics.stdev(tempi)
    figure = []

    fig1, ax1 = plt.subplots(figsize=(10, 6))
    ax1.hist(tempi, bins=30, color=BLU, edgecolor="white", alpha=0.85)
    ax1.axvline(media, color=ROSSO, linewidth=2.0, linestyle="--", label=f"Media: {media:.1f} h")
    ax1.axvline(media - dev, color=ROSSO, linewidth=1.0, linestyle=":", alpha=0.7)
    ax1.axvline(media + dev, color=ROSSO, linewidth=1.0, linestyle=":", alpha=0.7, label=f"Dev: ±{dev:.1f} h")
    _stile_plot(ax1, "Distribuzione tempi totali", "Tempo totale (ore)", "Numero simulazioni")
    ax1.legend(framealpha=0.9)
    fig1.tight_layout()
    figure.append(fig1)

    secca = [r for r in tutti_i_risultati if r["scenario_climatico"] == "secca"]
    umida = [r for r in tutti_i_risultati if r["scenario_climatico"] == "umida"]
    
    fig2, (ax2a, ax2b) = plt.subplots(1, 2, figsize=(12, 6))
    fig2.suptitle("Confronto climatica", fontweight="bold")
    
    for ax, estrazione, titolo, unita in [
        (ax2a, lambda x: [r["t_totale"] for r in x], "Tempo medio", "h"),
        (ax2b, lambda x: [calcola_perdite_totali(r) for r in x], "Perdite medie", "t")
    ]:
        vals_s, vals_u = estrazione(secca), estrazione(umida)
        medie = [statistics.mean(v) if v else 0 for v in (vals_s, vals_u)]
        devs = [statistics.stdev(v) if len(v) > 1 else 0 for v in (vals_s, vals_u)]
        
        bars = ax.bar(["Secca", "Umida"], medie, color=[ARANCIO, BLU], yerr=devs, capsize=6)
        for bar, val in zip(bars, medie):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f"{val:.1f} {unita}", ha="center", va="bottom", fontweight="bold")
        _stile_plot(ax, titolo, ylabel=unita)
    fig2.tight_layout()
    figure.append(fig2)

    return figure

def salva_grafici_pdf(tutti_i_risultati, nome_file="analisi_grafici.pdf"):
    with PdfPages(nome_file) as pdf:
        for fig in crea_figure_analisi(tutti_i_risultati):
            pdf.savefig(fig, bbox_inches="tight", dpi=150)
            plt.close(fig)

def esporta_excel_analisi(tutti_i_risultati, nome_file="analisi_1000_simulazioni.xlsx"):
    tempi_totali = [r["t_totale"] for r in tutti_i_risultati]
    perdite_totali = [calcola_perdite_totali(r) for r in tutti_i_risultati]

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Statistiche Generali"
    _riga_titolo(ws1, "STATISTICHE GENERALI - 1000 SIMULAZIONI", 2)
    _riga_intestazione(ws1, ["Indicatore", "Valore"])
    stat = [
        ("N. simulazioni", len(tutti_i_risultati)),
        ("Tempo medio (h)", round(statistics.mean(tempi_totali), 2)),
        ("Tempo minimo (h)", round(min(tempi_totali), 2)),
        ("Tempo massimo (h)", round(max(tempi_totali), 2)),
        ("Dev. standard (h)", round(statistics.stdev(tempi_totali), 2)),
        ("Perdite medie (t)", round(statistics.mean(perdite_totali), 2)),
        ("Perdite minime (t)", round(min(perdite_totali), 2)),
        ("Perdite massime (t)", round(max(perdite_totali), 2)),
    ]
    for i, (label, val) in enumerate(stat):
        _riga_dati(ws1, [label, val], pari=(i % 2 == 0))
    _adatta_colonne(ws1)

    ws2 = wb.create_sheet("Attese per Fase")
    fasi_da_analizzare = {
        "pulizia": "Pulizia", "essiccazione": "Essiccazione",
        "tostatura": "Tostatura (solo Soia)", "confez_A": "Confezionamento A",
        "macinazione": "Macinazione", "confez_B": "Confezionamento B",
    }
    attesa_media_per_fase = {}
    dati_fasi = []
    
    for chiave, nome in fasi_da_analizzare.items():
        attese, durate = [], []
        for r in tutti_i_risultati:
            if chiave == "tostatura":
                attese.append(r["fasi"]["soia"]["tostatura"]["t_attesa"])
                durate.append(r["fasi"]["soia"]["tostatura"]["tempo_ore"])
            else:
                for c in COLTURE:
                    attese.append(r["fasi"][c][chiave]["t_attesa"])
                    durate.append(r["fasi"][c][chiave]["tempo_ore"])
                    
        media_att = round(statistics.mean(attese), 2)
        attesa_media_per_fase[nome] = media_att
        dati_fasi.append([nome, media_att, round(max(attese), 2), round(statistics.mean(durate), 2), round(max(durate), 2)])

    collo = max(attesa_media_per_fase, key=attesa_media_per_fase.get)

    _riga_titolo(ws2, "ATTESE E DURATE MEDIE PER FASE", 6)
    _riga_intestazione(ws2, ["Fase", "Attesa media (h)", "Attesa max (h)", "Durata media (h)", "Durata max (h)", "Collo di bottiglia"])
    for i, riga in enumerate(dati_fasi):
        is_collo = riga[0] == collo
        _riga_dati(ws2, riga + ["<-- COLLO DI BOTTIGLIA" if is_collo else ""], pari=(i % 2 == 0), evidenzia=is_collo)
    _adatta_colonne(ws2)

    ws3 = wb.create_sheet("Secca vs Umida")
    _riga_titolo(ws3, "CONFRONTO ANNATA SECCA vs ANNATA UMIDA", 4)
    secca = [r for r in tutti_i_risultati if r["scenario_climatico"] == "secca"]
    umida = [r for r in tutti_i_risultati if r["scenario_climatico"] == "umida"]

    if not secca or not umida:
        ws3.append([f"Dati insufficienti: {len(secca)} secca, {len(umida)} umida."])
    else:
        _riga_intestazione(ws3, ["Indicatore", "Annata Secca", "Annata Umida", "Differenza"])
        ts, tu = [r["t_totale"] for r in secca], [r["t_totale"] for r in umida]
        ps, pu = [calcola_perdite_totali(r) for r in secca], [calcola_perdite_totali(r) for r in umida]
        
        ms, mu = round(statistics.mean(ts), 2), round(statistics.mean(tu), 2)
        ps_m, pu_m = round(statistics.mean(ps), 2), round(statistics.mean(pu), 2)
        
        confronto = [
            ("N. simulazioni", len(secca), len(umida), "-"),
            ("Tempo medio (h)", ms, mu, round(mu - ms, 2)),
            ("Dev. standard (h)", round(statistics.stdev(ts), 2) if len(ts) > 1 else "n/d", round(statistics.stdev(tu), 2) if len(tu) > 1 else "n/d", "-"),
            ("Tempo minimo (h)", round(min(ts), 2), round(min(tu), 2), "-"),
            ("Tempo massimo (h)", round(max(ts), 2), round(max(tu), 2), "-"),
            ("Perdite medie (t)", ps_m, pu_m, round(pu_m - ps_m, 2)),
            ("Perdite minime (t)", round(min(ps), 2), round(min(pu), 2), "-"),
            ("Perdite massime (t)", round(max(ps), 2), round(max(pu), 2), "-"),
        ]
        for i, riga in enumerate(confronto):
            _riga_dati(ws3, list(riga), pari=(i % 2 == 0))
    _adatta_colonne(ws3)

    ws4 = wb.create_sheet("Caso Migliore e Peggiore")
    _riga_titolo(ws4, "CASO MIGLIORE E CASO PEGGIORE", 3)
    
    migliore = min(tutti_i_risultati, key=lambda r: r["t_totale"])
    peggiore = max(tutti_i_risultati, key=lambda r: r["t_totale"])

    for etichetta, sim, colore in [("CASO MIGLIORE  -  tempo minimo", migliore, "E8F5E9"), ("CASO PEGGIORE  -  tempo massimo", peggiore, "FFEBEE")]:
        _riga_vuota(ws4, 3)
        _scrivi_riga(ws4, [etichetta], _font(True, 11), bg_color=colore, align="center", merge_cols=3)
        _riga_intestazione(ws4, ["Parametro", "Valore", ""])
        
        righe_sim = [
            ("Tempo totale (ore)", sim["t_totale"], ""),
            ("Tempo totale (giorni lav.)", sim["t_totale_gg"], ""),
            ("Perdite totali (t)", calcola_perdite_totali(sim), ""),
            ("Scenario raccolta", sim["scenario_raccolta"].capitalize(), ""),
            ("Annata climatica", sim["scenario_climatico"].capitalize(), ""),
        ]
        for c_nome in COLTURE:
            righe_sim.append((f"Resa {c_nome.capitalize()}", f"{sim['scenario_resa'][c_nome].capitalize()}  ({sim['quantita'][c_nome]:.2f} t)", ""))
            
        for i, riga in enumerate(righe_sim):
            _riga_dati(ws4, list(riga), pari=(i % 2 == 0))

    _adatta_colonne(ws4)
    wb.save(nome_file)